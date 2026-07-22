# importar_excel.py
# Lê a(s) Matriz(es) de Recolha de Informação sobre ONGs (.xlsx) e popula a base SQLite.
#
# Em vez de assumir posições fixas de colunas, este importador lê o cabeçalho
# de cada folha (a linha com "Designação da Organização" + a linha de
# sub-cabeçalhos por baixo) e identifica automaticamente a que campo cada
# coluna corresponde. Isto permite suportar formatos diferentes de matriz
# (ex.: o formato "clássico" com uma coluna "Distritos Beneficiários", ou o
# formato mais recente com "Cobertura Geográfica" dividida em "Província" e
# "Distrito/Município", "Programas" separado de "Projectos", e "Coordenadas
# UTM" X/Y) sem precisar de ajustar o código cada vez que aparece uma
# variante nova.
#
# A província de cada registo é determinada, por ordem de prioridade:
#   1. coluna explícita "Província" na própria linha (se existir nesse formato)
#   2. nome do ficheiro (mais fiável do que o cabeçalho interno)
#   3. célula "Província de: X" no cabeçalho da folha
#
# Uso:
#   python importar_excel.py                        -> importa todos os .xlsx em data/
#   python importar_excel.py caminho/ficheiro.xlsx   -> importa um único ficheiro
#   python importar_excel.py caminho/pasta           -> importa todos os .xlsx dessa pasta

import sys
import os
import glob
import openpyxl

from database import get_connection, init_db
from tagging import (
    normalizar, detetar_provincia, detetar_distritos, detetar_areas, detetar_ramificacoes,
    mapear_distrito_por_nome_aproximado, ESTRUTURA_AREAS_TEMATICAS,
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PASTA_DADOS_PADRAO = os.path.join(BASE_DIR, 'data')

LINHAS_PROCURAR_CABECALHO = 30
COL_NUMERO = 0  # "No." é sempre a primeira coluna em todos os formatos vistos


def limpar(valor):
    if valor is None:
        return None
    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)
    valor = str(valor).strip()
    return valor if valor else None


def localizar_linha_cabecalho(sheet):
    """Procura a linha com 'Designação da Organização' — é o âncora mais
    estável entre todos os formatos de matriz vistos até agora."""
    for r in range(1, min(LINHAS_PROCURAR_CABECALHO, sheet.max_row) + 1):
        for c in range(1, 4):
            valor = sheet.cell(row=r, column=c).value
            if valor and 'designacao' in normalizar(valor):
                return r
    return None


def localizar_texto_provincia_de(sheet, ate_linha):
    """Procura a célula 'Província de: X' (com dois pontos) acima do cabeçalho."""
    for r in range(1, ate_linha):
        for c in range(1, 4):
            valor = sheet.cell(row=r, column=c).value
            if valor and 'provincia de:' in normalizar(valor):
                return valor
    return None


def construir_mapa_colunas(sheet, linha_principal, linha_sub):
    """Lê as duas linhas de cabeçalho e devolve um dict campo_logico -> índice
    de coluna (0-based). Campos não encontrados simplesmente não aparecem no
    mapa (o valor fica None na importação, em vez de dar erro)."""
    max_col = sheet.max_column
    principal = [sheet.cell(row=linha_principal, column=c).value for c in range(1, max_col + 1)]
    sub = [sheet.cell(row=linha_sub, column=c).value for c in range(1, max_col + 1)]

    # cabeçalhos com células fundidas só têm texto na 1ª célula — propaga-se
    # esse texto para a direita até ao próximo cabeçalho não vazio
    principal_ff = []
    ultimo = None
    for v in principal:
        if v:
            ultimo = v
        principal_ff.append(ultimo)

    mapa = {}
    for idx in range(max_col):
        if idx == COL_NUMERO:
            continue
        texto_sub = normalizar(sub[idx]) if sub[idx] else ''
        texto_principal = normalizar(principal_ff[idx]) if principal_ff[idx] else ''

        if texto_sub:
            if texto_sub == 'ong':
                mapa['natureza_ong'] = idx
            elif 'osc' in texto_sub:
                mapa['natureza_osc'] = idx
            elif 'fundac' in texto_sub:
                mapa['natureza_fundacao'] = idx
            elif texto_sub == 'nacional':
                mapa['origem_nacional'] = idx
            elif texto_sub in ('internacional', 'estrangeira'):
                mapa['origem_internacional'] = idx
            elif texto_sub == 'x':
                mapa['coordenada_x'] = idx
            elif texto_sub == 'y':
                mapa['coordenada_y'] = idx
            elif texto_sub == 'telefone':
                mapa['telefone'] = idx
            elif 'mail' in texto_sub:
                mapa['email'] = idx
            elif texto_sub == 'provincia':
                mapa['provincia_coluna'] = idx
            elif 'distrito' in texto_sub or 'municipio' in texto_sub:
                mapa['distritos_texto'] = idx
            elif texto_sub.startswith('programa'):
                mapa['programas'] = idx
            elif texto_sub.startswith('projec') or texto_sub.startswith('projet'):
                mapa['projetos_em_curso'] = idx
        elif texto_principal:
            if 'designacao' in texto_principal:
                mapa['nome'] = idx
            elif 'area' in texto_principal and ('intervenc' in texto_principal or 'tematic' in texto_principal):
                mapa['area_intervencao'] = idx
            elif 'distrito' in texto_principal:
                mapa['distritos_texto'] = idx
            elif 'projec' in texto_principal or 'projet' in texto_principal:
                mapa['projetos_em_curso'] = idx
            elif 'financiamento' in texto_principal:
                mapa['fonte_financiamento'] = idx
            elif 'perfil' in texto_principal:
                mapa['perfil_grupo_alvo'] = idx
            elif 'ponto focal' in texto_principal or ('pessoa' in texto_principal and 'contacto' in texto_principal):
                mapa['pessoa_contacto'] = idx

    return mapa


def valor(row, mapa, campo):
    idx = mapa.get(campo)
    if idx is None or idx >= len(row):
        return None
    return limpar(row[idx])


def montar_natureza(row, mapa):
    partes = []
    if valor(row, mapa, 'natureza_ong'):
        partes.append('ONG')
    if valor(row, mapa, 'natureza_osc'):
        partes.append('OSC')
    if valor(row, mapa, 'natureza_fundacao'):
        partes.append('Fundação')
    return ', '.join(partes) if partes else None


def montar_origem(row, mapa):
    nacional = valor(row, mapa, 'origem_nacional')
    internacional = valor(row, mapa, 'origem_internacional')
    if nacional and internacional:
        return 'Nacional e Estrangeira'
    if nacional:
        return 'Nacional'
    if internacional:
        return 'Estrangeira'
    return None


SUB_PARA_AREA = {sub: area for area, subs in ESTRUTURA_AREAS_TEMATICAS.items() for sub in subs}


def obter_ou_criar_id(conn, tabela, nome, provincia=None):
    cur = conn.execute(f'SELECT id FROM {tabela} WHERE nome = ?', (nome,))
    linha = cur.fetchone()
    if linha:
        return linha['id']
    if provincia and tabela == 'distritos':
        cur = conn.execute('INSERT INTO distritos (nome, provincia) VALUES (?, ?)', (nome, provincia))
    else:
        cur = conn.execute(f'INSERT INTO {tabela} (nome) VALUES (?)', (nome,))
    return cur.lastrowid


def obter_ou_criar_ramificacao_id(conn, nome, area_id):
    cur = conn.execute('SELECT id FROM ramificacoes WHERE nome = ?', (nome,))
    linha = cur.fetchone()
    if linha:
        return linha['id']
    cur = conn.execute('INSERT INTO ramificacoes (nome, area_id) VALUES (?, ?)', (nome, area_id))
    return cur.lastrowid


def determinar_distrito_da_folha(nome_folha, provincia):
    return mapear_distrito_por_nome_aproximado(nome_folha, provincia)


def importar_folha(conn, sheet, nome_ficheiro, provincias_ja_limpas, provincia_do_ficheiro=None):
    linha_principal = localizar_linha_cabecalho(sheet)
    if not linha_principal:
        return {}, 0  # folha sem cabeçalho reconhecível — ignorada silenciosamente

    linha_sub = linha_principal + 1
    primeira_linha_dados = linha_principal + 2
    mapa = construir_mapa_colunas(sheet, linha_principal, linha_sub)

    texto_provincia_celula = localizar_texto_provincia_de(sheet, linha_principal)
    provincia_da_celula = detetar_provincia(texto_provincia_celula) if texto_provincia_celula else None
    provincia_da_folha = provincia_do_ficheiro or provincia_da_celula

    if provincia_da_celula and provincia_do_ficheiro and provincia_da_celula != provincia_do_ficheiro:
        print(f'  AVISO: "{nome_ficheiro}" (folha "{sheet.title}") - o nome do ficheiro indica '
              f'"{provincia_do_ficheiro}" mas a celula interna diz "{provincia_da_celula}". '
              f'A usar "{provincia_do_ficheiro}" (nome do ficheiro). Revise a matriz de origem se nao for o esperado.')

    distrito_da_folha = determinar_distrito_da_folha(sheet.title, provincia_da_folha) if provincia_da_folha else None
    avisou_inconsistencia = False
    resumo = {}

    for row in sheet.iter_rows(min_row=primeira_linha_dados, values_only=True):
        nome = valor(row, mapa, 'nome') if mapa.get('nome') is not None else None
        if not nome or normalizar(nome) == 'legenda':
            continue
        if len(row) <= COL_NUMERO or not limpar(row[COL_NUMERO]):
            continue

        # provincia: prioridade para a coluna explicita da propria linha (se existir),
        # depois o nome do ficheiro, depois a celula "Provincia de:" da folha
        provincia_coluna_texto = valor(row, mapa, 'provincia_coluna')
        provincia_da_linha = (detetar_provincia(provincia_coluna_texto) if provincia_coluna_texto else None) \
            or provincia_da_folha
        if not provincia_da_linha:
            continue  # sem forma de saber a que provincia pertence - ignora a linha

        if provincia_da_linha not in provincias_ja_limpas:
            conn.execute('DELETE FROM ongs WHERE provincia = ?', (provincia_da_linha,))
            provincias_ja_limpas.add(provincia_da_linha)

        numero_original = row[COL_NUMERO]
        natureza = montar_natureza(row, mapa)
        origem = montar_origem(row, mapa)
        area_texto = valor(row, mapa, 'area_intervencao')
        programas = valor(row, mapa, 'programas')
        projetos = valor(row, mapa, 'projetos_em_curso')
        financiamento = valor(row, mapa, 'fonte_financiamento')
        distritos_texto = valor(row, mapa, 'distritos_texto') or distrito_da_folha
        perfil = valor(row, mapa, 'perfil_grupo_alvo')
        pessoa = valor(row, mapa, 'pessoa_contacto')
        telefone = valor(row, mapa, 'telefone')
        email = valor(row, mapa, 'email')
        coord_x = valor(row, mapa, 'coordenada_x')
        coord_y = valor(row, mapa, 'coordenada_y')

        cur = conn.execute('''
            INSERT INTO ongs (numero_original, nome, natureza, origem, provincia,
                               area_intervencao, programas, projetos_em_curso, fonte_financiamento,
                               distritos_texto, perfil_grupo_alvo, pessoa_contacto, telefone, email,
                               coordenada_x, coordenada_y)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (numero_original, nome, natureza, origem, provincia_da_linha, area_texto, programas,
              projetos, financiamento, distritos_texto, perfil, pessoa, telefone, email, coord_x, coord_y))
        ong_id = cur.lastrowid

        distritos_detetados = set(detetar_distritos(distritos_texto, provincia_da_linha))
        if distrito_da_folha and distrito_da_folha not in distritos_detetados:
            if not avisou_inconsistencia:
                print(f'  AVISO: folha "{sheet.title}" ({provincia_da_linha}) - coluna de distrito '
                      f'não corresponde ao distrito da folha em algumas linhas (ex.: "{distritos_texto}"). '
                      f'A registar com ambos os distritos; reveja a matriz de origem quando possível.')
                avisou_inconsistencia = True
            distritos_detetados.add(distrito_da_folha)

        for distrito in sorted(distritos_detetados):
            distrito_id = obter_ou_criar_id(conn, 'distritos', distrito, provincia_da_linha)
            conn.execute('INSERT OR IGNORE INTO ong_distritos (ong_id, distrito_id) VALUES (?, ?)',
                         (ong_id, distrito_id))

        for area in detetar_areas(area_texto):
            area_id = obter_ou_criar_id(conn, 'areas', area)
            conn.execute('INSERT OR IGNORE INTO ong_areas (ong_id, area_id) VALUES (?, ?)',
                         (ong_id, area_id))

        for ramificacao in detetar_ramificacoes(area_texto):
            area_da_ramificacao = SUB_PARA_AREA.get(ramificacao)
            if not area_da_ramificacao:
                continue
            area_id = obter_ou_criar_id(conn, 'areas', area_da_ramificacao)
            ramificacao_id = obter_ou_criar_ramificacao_id(conn, ramificacao, area_id)
            conn.execute('INSERT OR IGNORE INTO ong_ramificacoes (ong_id, ramificacao_id) VALUES (?, ?)',
                         (ong_id, ramificacao_id))

        resumo[provincia_da_linha] = resumo.get(provincia_da_linha, 0) + 1

    return resumo, sum(resumo.values())


def importar_ficheiro(conn, caminho_xlsx, provincias_ja_limpas):
    wb = openpyxl.load_workbook(caminho_xlsx, data_only=True)
    nome_ficheiro = os.path.basename(caminho_xlsx)

    # tenta identificar a provincia pelo nome do ficheiro (mais fiavel do que
    # a celula "Provincia de:" dentro da folha, que por vezes vem errada)
    provincia_do_ficheiro = detetar_provincia(os.path.splitext(nome_ficheiro)[0])

    resumo_total = {}
    for sheet in wb.worksheets:
        resumo_folha, _ = importar_folha(conn, sheet, nome_ficheiro, provincias_ja_limpas, provincia_do_ficheiro)
        for provincia, total in resumo_folha.items():
            resumo_total[provincia] = resumo_total.get(provincia, 0) + total
        conn.commit()

    if not resumo_total:
        print(f'  AVISO: nenhum registo importado de "{nome_ficheiro}" (cabeçalho não reconhecido ou ficheiro vazio).')
        return 0

    total_geral = 0
    for provincia, total in resumo_total.items():
        print(f'  OK {provincia}: {total} registos importados ({nome_ficheiro})')
        total_geral += total
    return total_geral


def listar_ficheiros(caminho):
    if os.path.isdir(caminho):
        return sorted(glob.glob(os.path.join(caminho, '*.xlsx')))
    if os.path.isfile(caminho):
        return [caminho]
    return []


def importar(caminho):
    init_db()
    ficheiros = listar_ficheiros(caminho)
    if not ficheiros:
        print(f'Nenhum ficheiro .xlsx encontrado em: {caminho}')
        sys.exit(1)

    conn = get_connection()
    provincias_ja_limpas = set()
    total_geral = 0
    print(f'A importar {len(ficheiros)} ficheiro(s)...')
    for ficheiro in ficheiros:
        total_geral += importar_ficheiro(conn, ficheiro, provincias_ja_limpas)
    conn.close()
    print(f'\nImportacao concluida: {total_geral} registos no total.')


if __name__ == '__main__':
    caminho = sys.argv[1] if len(sys.argv) > 1 else PASTA_DADOS_PADRAO
    if not os.path.exists(caminho):
        print(f'Caminho nao encontrado: {caminho}')
        sys.exit(1)
    importar(caminho)
